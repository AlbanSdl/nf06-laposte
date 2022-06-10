from ctypes import POINTER, Structure, CDLL, byref, c_float, c_int
import openpyxl
import networkx as nx
import matplotlib.pyplot as plt


class Car(Structure):
    """
    Python equivalent for the vehicule structure (in C)
    Linked with C through Ctypes. The _fields_ property defines the fields
    of the structure so that ctypes bindings are working properly (with the proper c types)

    This class contains 4 fields, defined as:
    - capacity: the capacity (in customers) of the vehicle
    - battery: the autonomy of the vehicle (in km)
    - speedkmh: the speed of the vehicle (in km/h)
    - rechargetps: the refuel time of the vehicle (in hours)
    """
    _fields_ = ("capacity", c_int), ("battery", c_int), ("speedkmh", c_float), ("rechargetps", c_float)

class Node(Structure):
    """
    Python equivalent for the node structure (in C)
    Linked with C through Ctypes (as we're inheriting the Structure class). The _fields_ property defines the fields
    of the structure so that ctypes bindings are working properly (with the proper c types)

    This class contains 2 fields, defined as:
    - id: the id of the node
    - distances: an array containing the distance to the other nodes (in km). The array is indexed by the id of the nodes
    """
    _fields_ = ("id", c_int), ("distances", POINTER(c_float))

class RoutedVehicle(Structure):
    """
    Python equivalent for the routed vehicle structure (in C)
    Linked with C through Ctypes (as we're inheriting the Structure class). The _fields_ property defines the fields
    of the structure so that ctypes bindings are working properly (with the proper c types)

    This class contains 5 fields, defined as:
    - nodes: an array containing the nodes of the route (the array only contains the ids of the nodes). This array does not contain the starting point (which is the warehouse)
    - nodeCount: the length of the nodes array
    - deliveries: the number of deliveres of the vehicle (ie. nodeCount - amount of travels to the warehouse)
    - distance: the distance of the route (in km) for this vehicle
    - time: the time needed by the vehicle to complete the route (in hours)
    """
    _fields_ = ("nodes", POINTER(c_int)), ("nodeCount", c_int), ("deliveries", c_int), ("distance", c_float), ("time", c_float)


def computePath(nodes, cars):
    """
    This function calls our C library to compute the routes

    Parameters
    ----------
    nodes:  list[Node]
            a python list of the c nodes (of type Node)
    cars:   list[Car]
            a python list of cars (of type Car)
    
    Returns a ctypes array of the routedVehicles (of type RoutedVehicle)
    """
    # Load library
    lib = CDLL('build/libevrp.dll')
    computePath = lib.compute_path
    # Initialize the output of the c function. The output will be retrieved
    # using a pointer (that's why we call `byref`) along with `computePath`
    result = (RoutedVehicle * len(cars))()
    computePath(
        (Node * len(nodes))(*nodes),
        c_int(len(nodes)),
        (Car * len(cars))(*cars),
        c_int(len(cars)),
        byref(result)
    )
    return result


def readFromFile(path):
    """
    Reads data (both nodes and cars) from an excel file, computes distances between nodes

    Parameters
    ----------
    path:   str
            the path to the excel file

    Returns a tuple containing:
    - nodes:  list[Node]
            a python list of the c nodes (of type Node)
    - cars:   list[Car]
            a python list of cars (of type Car)
    """
    # Load the excel file
    wb_obj = openpyxl.load_workbook(path)
    # Get the first sheet
    node_sheet = wb_obj["Feuil1"]

    # Populates the node list using the excel file (we retrieve the coordinates and assign ids to the nodes)
    # The id is determined with the row of the entry in the excel file
    nodes = []
    for j in range(2, node_sheet.max_row + 1):
        nodes.append({
            "id": j - 2,
            "x": node_sheet.cell(row=j, column=1).value,
            "y": node_sheet.cell(row=j, column=2).value,
        })

    # We compute distances between each nodes and prepare data types to use when
    # calling the C library
    c_nodes = []
    for node in nodes:
        c_nodes.append(Node(id=node["id"], distances=(c_float * len(nodes))(
            *[((node["x"] - n["x"]) ** 2 + (node["y"] - n["y"]) ** 2) ** .5 for n in nodes]
        )))
    
    # Access the second sheet
    car_sheet = wb_obj["Feuil2"]
    cars = []
    # Populates the car list using the excel file (we retrieve the capacity, battery, speed and refueling speed)
    for i in range(2, car_sheet.max_row + 1):
        cars.append(Car(
            capacity=car_sheet.cell(row=i, column=1).value,
            battery=car_sheet.cell(row=i, column=2).value,
            speedkmh=car_sheet.cell(row=i, column=4).value,
            rechargetps=car_sheet.cell(row=i, column=3).value
        ))

    # Close excel file
    wb_obj.close()
    # Return the data
    return (nodes, c_nodes, cars)


def displayResult(nodes, routedVehicles):
    """
    Displays the result of the computation

    Parameters
    ----------
    nodes:          list[Node]
                    a python list of the nodes (with their coordinates, thus NOT the c nodes)
    routedVehicles: list[RoutedVehicle]
                    a python list of the c routedVehicles (of type RoutedVehicle)
    """
    ## Creating the graph
    G=nx.Graph()

    # Add all nodes to the graph
    for node in nodes:
        G.add_node(node["id"], pos=(node["x"], node["y"]))

    # Define colors to use for the vehicles (the first vehicle will use red, etc)
    colors = ["red", "blue", "green", "yellow", "orange", "purple", "pink", "brown", "black", "grey"]

    # Index of the color to use for the vehicle
    nb = 0
    # We add the paths to the graph and use a different color for each vehicle
    for i in routedVehicles:
        for j in range(0, i.nodeCount - 1):
            if j == 0:
                G.add_edge(0, i.nodes[j], color=colors[nb % len(colors)])
            G.add_edge(i.nodes[j], i.nodes[j + 1], color=colors[nb % len(colors)])
        nb += 1

    # We scale our graph properly
    pos=nx.get_node_attributes(G, 'pos')
    # We compute colors for the edges
    color = [G[u][v]['color'] for u,v in G.edges()]
    # We render the graph
    nx.draw(G, pos, with_labels=True, edge_color=color)
    # We display the graph
    plt.show()


# We read the data from the excel file and store data nodes and the cars
# The nodes have coordinates and they will be used in python code only (for the rendering)
# c_nodes will only be used in C code, and contains the distances between each nodes
# cars will be used in both python and C code
(nodes, c_nodes, cars) = readFromFile("lieux.xlsx")
## The routes computed by the c library
routes = computePath(c_nodes, cars)

# We display the stats for all the routes (in console)
for route in routes:
    print(f"{route.deliveries} deliveries in {route.time} hours (after {route.distance} km)")

# Display graph at the screen
displayResult(nodes, routes)